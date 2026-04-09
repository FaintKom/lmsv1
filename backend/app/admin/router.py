import csv
import io
import uuid

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm.attributes import flag_modified

from app.admin.schemas import DashboardStats, DetailedAnalytics
from app.admin.service import get_dashboard_stats, get_detailed_analytics
from app.auth.dependencies import require_role
from app.auth.models import Organization, User, UserRole
from app.auth.schemas import UserResponse
from app.common.exceptions import NotFoundError
from app.db.session import get_db

router = APIRouter()


# ─── Organization Management ──────────────────────────────────────────


class OrgUpdate(BaseModel):
    name: str | None = None
    is_active: bool | None = None
    settings: dict | None = None


@router.get("/organizations")
async def list_organizations_endpoint(
    admin: User = Depends(require_role(UserRole.admin)),
    db: AsyncSession = Depends(get_db),
):
    """List organizations. Super admin sees all, regular admin sees own only."""
    query = select(Organization)
    if admin.role != UserRole.super_admin:
        query = query.where(Organization.id == admin.org_id)
    result = await db.execute(query.order_by(Organization.name))
    orgs = result.scalars().all()
    return [
        {
            "id": str(o.id),
            "name": o.name,
            "slug": o.slug,
            "is_active": o.is_active,
            "settings": o.settings or {},
            "created_at": str(o.created_at) if o.created_at else None,
        }
        for o in orgs
    ]


@router.get("/organizations/{org_id}")
async def get_organization_endpoint(
    org_id: uuid.UUID,
    admin: User = Depends(require_role(UserRole.admin)),
    db: AsyncSession = Depends(get_db),
):
    """Get single organization."""
    if admin.role != UserRole.super_admin and admin.org_id != org_id:
        raise NotFoundError("Organization not found")
    result = await db.execute(select(Organization).where(Organization.id == org_id))
    org = result.scalar_one_or_none()
    if not org:
        raise NotFoundError("Organization not found")
    return {
        "id": str(org.id),
        "name": org.name,
        "slug": org.slug,
        "is_active": org.is_active,
        "settings": org.settings or {},
        "created_at": str(org.created_at) if org.created_at else None,
    }


@router.put("/organizations/{org_id}")
async def update_organization_endpoint(
    org_id: uuid.UUID,
    data: OrgUpdate,
    admin: User = Depends(require_role(UserRole.admin)),
    db: AsyncSession = Depends(get_db),
):
    """Update organization. Admin can update own org, super admin can update any."""
    if admin.role != UserRole.super_admin and admin.org_id != org_id:
        raise NotFoundError("Organization not found")
    result = await db.execute(select(Organization).where(Organization.id == org_id))
    org = result.scalar_one_or_none()
    if not org:
        raise NotFoundError("Organization not found")
    if data.name is not None:
        org.name = data.name
    if data.is_active is not None and admin.role == UserRole.super_admin:
        org.is_active = data.is_active
    if data.settings is not None:
        # Merge with existing settings — must flag_modified for JSONB
        current = dict(org.settings or {})
        current.update(data.settings)
        org.settings = current
        flag_modified(org, "settings")
    await db.commit()
    return {
        "id": str(org.id),
        "name": org.name,
        "slug": org.slug,
        "is_active": org.is_active,
        "settings": org.settings or {},
    }


@router.delete("/organizations/{org_id}")
async def delete_organization_endpoint(
    org_id: uuid.UUID,
    admin: User = Depends(require_role(UserRole.admin)),
    db: AsyncSession = Depends(get_db),
):
    """Delete organization. Super admin only."""
    if admin.role != UserRole.super_admin:
        from fastapi import HTTPException
        raise HTTPException(403, "Only super admin can delete organizations")
    result = await db.execute(select(Organization).where(Organization.id == org_id))
    org = result.scalar_one_or_none()
    if not org:
        raise NotFoundError("Organization not found")
    # Prevent deleting own org
    if org.id == admin.org_id:
        from fastapi import HTTPException
        raise HTTPException(400, "Cannot delete your own organization")
    await db.delete(org)
    await db.commit()
    return {"status": "ok"}


def _user_org_filter(admin: User):
    """Return org filter for User queries — empty for super_admin."""
    if admin.role == UserRole.super_admin:
        return []
    return [User.org_id == admin.org_id]


@router.get("/dashboard", response_model=DashboardStats)
async def dashboard_endpoint(
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
    db: AsyncSession = Depends(get_db),
):
    return await get_dashboard_stats(db, user)


@router.get("/teacher-stats")
async def teacher_stats_endpoint(
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
    db: AsyncSession = Depends(get_db),
):
    """Teacher-specific dashboard stats."""
    from sqlalchemy import func

    from app.assignments.models import Assignment, AssignmentStatus, AssignmentSubmission
    from app.courses.models import Course
    from app.progress.models import Enrollment

    # My courses count
    course_q = select(func.count(Course.id)).where(Course.org_id == user.org_id)
    if user.role == UserRole.teacher:
        course_q = course_q.where(Course.teacher_id == user.id)
    my_courses = (await db.execute(course_q)).scalar() or 0

    # My students (enrolled in my courses)
    student_q = (
        select(func.count(func.distinct(Enrollment.student_id)))
        .join(Course, Enrollment.course_id == Course.id)
        .where(Course.org_id == user.org_id)
    )
    if user.role == UserRole.teacher:
        student_q = student_q.where(Course.teacher_id == user.id)
    my_students = (await db.execute(student_q)).scalar() or 0

    # Ungraded submissions
    ungraded_q = (
        select(func.count(AssignmentSubmission.id))
        .join(Assignment, AssignmentSubmission.assignment_id == Assignment.id)
        .where(
            AssignmentSubmission.status.in_([AssignmentStatus.submitted, AssignmentStatus.late]),
            Assignment.org_id == user.org_id,
        )
    )
    if user.role == UserRole.teacher:
        ungraded_q = ungraded_q.where(Assignment.created_by == user.id)
    to_review = (await db.execute(ungraded_q)).scalar() or 0

    # Average score across graded submissions
    avg_q = (
        select(func.avg(AssignmentSubmission.score))
        .join(Assignment, AssignmentSubmission.assignment_id == Assignment.id)
        .where(
            AssignmentSubmission.status == AssignmentStatus.graded,
            Assignment.org_id == user.org_id,
        )
    )
    if user.role == UserRole.teacher:
        avg_q = avg_q.where(Assignment.created_by == user.id)
    avg_score = (await db.execute(avg_q)).scalar()
    avg_score = round(float(avg_score), 1) if avg_score else 0

    # Recent submissions
    recent_q = (
        select(
            AssignmentSubmission,
            Assignment.title.label("assignment_title"),
            User.full_name.label("student_name"),
        )
        .join(Assignment, AssignmentSubmission.assignment_id == Assignment.id)
        .join(User, AssignmentSubmission.student_id == User.id)
        .where(Assignment.org_id == user.org_id)
    )
    if user.role == UserRole.teacher:
        recent_q = recent_q.where(Assignment.created_by == user.id)
    recent_q = recent_q.order_by(AssignmentSubmission.submitted_at.desc()).limit(5)
    result = await db.execute(recent_q)
    recent = [
        {
            "id": str(sub.id),
            "assignment_title": at,
            "student_name": sn,
            "submitted_at": str(sub.submitted_at),
            "status": sub.status.value if hasattr(sub.status, "value") else sub.status,
            "score": float(sub.score) if sub.score is not None else None,
        }
        for sub, at, sn in result.all()
    ]

    return {
        "my_courses": my_courses,
        "my_students": my_students,
        "to_review": to_review,
        "avg_score": avg_score,
        "recent_submissions": recent,
    }


@router.get("/analytics/detailed", response_model=DetailedAnalytics)
async def detailed_analytics_endpoint(
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
    db: AsyncSession = Depends(get_db),
):
    return await get_detailed_analytics(db, user)


@router.get("/users", response_model=list[UserResponse])
async def list_users_endpoint(
    role: str | None = Query(None),
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
    db: AsyncSession = Depends(get_db),
):
    query = select(User).where(*_user_org_filter(user))
    if role:
        query = query.where(User.role == role)
    result = await db.execute(query.order_by(User.created_at.desc()))
    users = result.scalars().all()
    return [UserResponse.model_validate(u) for u in users]


class _UpdateUserBody(BaseModel):
    role: str | None = None
    is_active: bool | None = None
    is_methodist: bool | None = None
    org_id: str | None = None  # super_admin only


@router.put("/users/{user_id}", response_model=UserResponse)
async def update_user_endpoint(
    user_id: uuid.UUID,
    body: _UpdateUserBody,
    admin: User = Depends(require_role(UserRole.admin)),
    db: AsyncSession = Depends(get_db),
):
    query = select(User).where(User.id == user_id, *_user_org_filter(admin))
    result = await db.execute(query)
    target_user = result.scalar_one_or_none()
    if not target_user:
        raise NotFoundError("User not found")

    if body.role is not None:
        # Only super_admin can assign the admin role
        if body.role == UserRole.admin.value and admin.role != UserRole.super_admin:
            from fastapi import HTTPException as _HTTPException
            raise _HTTPException(403, "Only super admin can assign admin role")
        target_user.role = body.role
    if body.is_active is not None:
        target_user.is_active = body.is_active
    if body.is_methodist is not None:
        target_user.is_methodist = body.is_methodist
    if body.org_id is not None and admin.role == UserRole.super_admin:
        target_user.org_id = uuid.UUID(body.org_id)

    await db.flush()
    return UserResponse.model_validate(target_user)


@router.get("/organizations")
async def list_organizations(
    admin: User = Depends(require_role(UserRole.admin)),
    db: AsyncSession = Depends(get_db),
):
    """List all organizations. Super admin sees all, regular admin sees own."""
    from sqlalchemy import func

    query = select(Organization)
    if admin.role != UserRole.super_admin:
        query = query.where(Organization.id == admin.org_id)
    result = await db.execute(query.order_by(Organization.name))
    orgs = result.scalars().all()

    org_list = []
    for o in orgs:
        # Count users in this org
        count_result = await db.execute(
            select(func.count()).select_from(User).where(User.org_id == o.id)
        )
        user_count = count_result.scalar() or 0
        org_list.append({
            "id": str(o.id),
            "name": o.name,
            "slug": o.slug,
            "is_active": o.is_active,
            "user_count": user_count,
            "created_at": str(o.created_at) if hasattr(o, 'created_at') and o.created_at else None,
        })
    return org_list


@router.get("/courses")
async def list_courses_admin(
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
    db: AsyncSession = Depends(get_db),
):
    from app.courses.models import Course

    query = select(Course)
    if user.role != UserRole.super_admin:
        query = query.where(Course.org_id == user.org_id)
    result = await db.execute(query.order_by(Course.created_at.desc()))
    courses = result.scalars().all()
    return [
        {
            "id": str(c.id), "title": c.title, "slug": c.slug,
            "description": c.description, "status": c.status.value if hasattr(c.status, 'value') else c.status,
            "category": c.category, "org_id": str(c.org_id), "created_at": str(c.created_at),
            "is_template": getattr(c, 'is_template', False),
            "thumbnail_url": getattr(c, 'thumbnail_url', None),
            "source_course_id": str(c.source_course_id) if getattr(c, 'source_course_id', None) else None,
            "template_version": getattr(c, 'template_version', 0),
        }
        for c in courses
    ]


class _UpdateCourseAdminBody(BaseModel):
    org_id: str | None = None


@router.put("/courses/{course_id}")
async def update_course_admin(
    course_id: uuid.UUID,
    body: _UpdateCourseAdminBody,
    admin: User = Depends(require_role(UserRole.admin)),
    db: AsyncSession = Depends(get_db),
):
    """Super admin can change course org_id."""
    from app.courses.models import Course

    if admin.role != UserRole.super_admin:
        from fastapi import HTTPException
        raise HTTPException(403, "Only super admin can change course organization")

    result = await db.execute(select(Course).where(Course.id == course_id))
    course = result.scalar_one_or_none()
    if not course:
        raise NotFoundError("Course not found")

    if body.org_id is not None:
        course.org_id = uuid.UUID(body.org_id)

    await db.flush()
    return {"ok": True}


@router.delete("/users/{user_id}")
async def delete_user_endpoint(
    user_id: uuid.UUID,
    admin: User = Depends(require_role(UserRole.admin)),
    db: AsyncSession = Depends(get_db),
):
    from fastapi import HTTPException

    query = select(User).where(User.id == user_id, *_user_org_filter(admin))
    result = await db.execute(query)
    target = result.scalar_one_or_none()
    if not target:
        raise NotFoundError("User not found")
    if target.id == admin.id:
        raise HTTPException(400, "Cannot delete yourself")
    await db.delete(target)
    return {"ok": True}


@router.post("/enroll")
async def admin_enroll_endpoint(
    data: dict,
    admin: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
    db: AsyncSession = Depends(get_db),
):
    """Admin enrolls a student into a course."""
    from datetime import datetime, timezone

    from app.courses.models import Course
    from app.progress.models import Enrollment

    user_id = data.get("user_id")
    course_id = data.get("course_id")
    if not user_id or not course_id:
        from fastapi import HTTPException
        raise HTTPException(400, "user_id and course_id required")

    # Verify user (super_admin can enroll any user)
    query = select(User).where(User.id == user_id, *_user_org_filter(admin))
    result = await db.execute(query)
    target_user = result.scalar_one_or_none()
    if not target_user:
        raise NotFoundError("User not found")

    # Verify course
    course_q = select(Course).where(Course.id == course_id)
    if admin.role != UserRole.super_admin:
        course_q = course_q.where(Course.org_id == admin.org_id)
    result = await db.execute(course_q)
    course = result.scalar_one_or_none()
    if not course:
        raise NotFoundError("Course not found")

    # Cannot enroll into template courses
    if getattr(course, "is_template", False):
        from fastapi import HTTPException
        raise HTTPException(400, "Cannot enroll students into a template course. Copy the template first.")

    # Check not already enrolled
    result = await db.execute(
        select(Enrollment).where(
            Enrollment.course_id == course_id, Enrollment.student_id == user_id
        )
    )
    if result.scalar_one_or_none():
        return {"status": "already_enrolled"}

    enrollment = Enrollment(
        course_id=course_id,
        student_id=user_id,
        enrolled_at=datetime.now(timezone.utc),
    )
    db.add(enrollment)
    await db.flush()
    return {"status": "ok", "enrollment_id": str(enrollment.id)}


@router.post("/bulk-enroll")
async def admin_bulk_enroll_endpoint(
    data: dict,
    admin: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
    db: AsyncSession = Depends(get_db),
):
    """Bulk-import students from a CSV payload and enroll them in a course.

    Body shape (JSON):
        {
          "course_id": "<uuid>",
          "rows": [
            {"email": "...", "full_name": "...", "password": "..."},
            ...
          ],
          "default_password": "Welcome2026!"   # optional fallback
        }

    Per-row behaviour:
    - Validate email format.
    - If a user with that email already exists in the admin's org, reuse them.
    - If not, create a new student with role=student, auto-verified (since
      admin is vouching for them via bulk import), consent recorded.
    - Enroll in the course unless already enrolled.
    - Capture per-row errors without aborting the whole batch.

    Returns a summary:
        {
          "total": N,
          "created": X,
          "reused": Y,
          "enrolled": Z,
          "already_enrolled": W,
          "errors": [ {"row": i, "email": "...", "message": "..."}, ... ]
        }

    The front-end parses a CSV locally and POSTs JSON — that way the
    backend doesn't need to know about CSV dialect quirks and large
    files don't have to be streamed.
    """
    import re as _re
    from datetime import datetime, timezone

    from app.auth.models import UserRole as _UserRole
    from app.auth.security import hash_password
    from app.courses.models import Course
    from app.progress.models import Enrollment

    course_id_str = data.get("course_id")
    rows = data.get("rows")
    default_password = (data.get("default_password") or "").strip()

    if not course_id_str or not isinstance(rows, list):
        raise HTTPException(400, "course_id and rows[] are required")

    try:
        course_id = uuid.UUID(course_id_str)
    except ValueError:
        raise HTTPException(400, "Invalid course_id") from None

    # Verify course exists and belongs to admin's org
    course_q = select(Course).where(Course.id == course_id)
    if admin.role != _UserRole.super_admin:
        course_q = course_q.where(Course.org_id == admin.org_id)
    course = (await db.execute(course_q)).scalar_one_or_none()
    if not course:
        raise NotFoundError("Course not found")
    if getattr(course, "is_template", False):
        raise HTTPException(
            400,
            "Cannot bulk-enroll students into a template course. Copy the template first.",
        )

    EMAIL_RE = _re.compile(r"^[\w.+-]+@[\w-]+(\.[\w-]+)+$")  # noqa: N806

    total = len(rows)
    created = 0
    reused = 0
    enrolled = 0
    already_enrolled = 0
    errors: list[dict] = []

    for i, raw_row in enumerate(rows):
        row_num = i + 1  # 1-based for user-facing messages
        if not isinstance(raw_row, dict):
            errors.append({"row": row_num, "email": "", "message": "Row must be an object"})
            continue

        email = (raw_row.get("email") or "").strip().lower()
        full_name = (raw_row.get("full_name") or raw_row.get("name") or "").strip()
        password = (raw_row.get("password") or default_password).strip()

        if not email or not EMAIL_RE.match(email):
            errors.append({"row": row_num, "email": email, "message": "Invalid email"})
            continue
        if not full_name:
            errors.append({"row": row_num, "email": email, "message": "Missing full_name"})
            continue
        if not password or len(password) < 8:
            errors.append(
                {"row": row_num, "email": email, "message": "Password must be 8+ chars"}
            )
            continue

        try:
            # Look up existing user
            existing = (
                await db.execute(select(User).where(User.email == email))
            ).scalar_one_or_none()

            if existing:
                # Only reuse if they're in the admin's org (or admin is super)
                if (
                    admin.role != _UserRole.super_admin
                    and existing.org_id != admin.org_id
                ):
                    errors.append(
                        {
                            "row": row_num,
                            "email": email,
                            "message": "User exists in a different organization",
                        }
                    )
                    continue
                user_row = existing
                reused += 1
            else:
                user_row = User(
                    org_id=admin.org_id,
                    email=email,
                    hashed_password=hash_password(password),
                    full_name=full_name,
                    role=_UserRole.student,
                    is_active=True,
                    consent_accepted_at=datetime.now(timezone.utc),
                    privacy_policy_version="1.0",
                    # Bulk-imported students are vouched for by the admin —
                    # auto-verified, same policy as invite-link registration.
                    email_verified_at=datetime.now(timezone.utc),
                )
                db.add(user_row)
                await db.flush()
                created += 1

            # Enroll if not already
            existing_enroll = (
                await db.execute(
                    select(Enrollment).where(
                        Enrollment.course_id == course_id,
                        Enrollment.student_id == user_row.id,
                    )
                )
            ).scalar_one_or_none()
            if existing_enroll:
                already_enrolled += 1
            else:
                db.add(
                    Enrollment(
                        course_id=course_id,
                        student_id=user_row.id,
                        enrolled_at=datetime.now(timezone.utc),
                    )
                )
                await db.flush()
                enrolled += 1

        except Exception as e:
            errors.append({"row": row_num, "email": email, "message": str(e)[:200]})

    return {
        "total": total,
        "created": created,
        "reused": reused,
        "enrolled": enrolled,
        "already_enrolled": already_enrolled,
        "errors": errors,
    }


@router.get("/courses/{course_id}/students")
async def list_course_students(
    course_id: uuid.UUID,
    admin: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
    db: AsyncSession = Depends(get_db),
):
    """List all students enrolled in a course."""
    from app.progress.models import Enrollment

    query = (
        select(User, Enrollment)
        .join(Enrollment, Enrollment.student_id == User.id)
        .where(Enrollment.course_id == course_id)
    )
    if admin.role != UserRole.super_admin:
        query = query.where(User.org_id == admin.org_id)
    result = await db.execute(query.order_by(Enrollment.enrolled_at.desc()))
    rows = result.all()
    return [
        {
            "id": str(u.id),
            "full_name": u.full_name,
            "email": u.email,
            "role": u.role,
            "enrollment_id": str(e.id),
            "progress_percent": e.progress_percent,
            "enrolled_at": str(e.enrolled_at),
        }
        for u, e in rows
    ]


@router.delete("/enrollments/{enrollment_id}")
async def admin_unenroll_endpoint(
    enrollment_id: uuid.UUID,
    admin: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
    db: AsyncSession = Depends(get_db),
):
    """Admin removes a student enrollment."""
    from app.courses.models import Course
    from app.progress.models import Enrollment

    query = (
        select(Enrollment)
        .join(Course, Enrollment.course_id == Course.id)
        .where(Enrollment.id == enrollment_id)
    )
    if admin.role != UserRole.super_admin:
        query = query.where(Course.org_id == admin.org_id)
    result = await db.execute(query)
    enrollment = result.scalar_one_or_none()
    if not enrollment:
        raise NotFoundError("Enrollment not found")

    await db.delete(enrollment)
    await db.flush()
    return {"status": "ok"}


@router.post("/users", response_model=UserResponse)
async def create_user_endpoint(
    data: dict,
    admin: User = Depends(require_role(UserRole.admin)),
    db: AsyncSession = Depends(get_db),
):
    from app.auth.security import hash_password

    role = data.get("role", "student")
    # Only super_admin can create admin users
    if role == UserRole.admin.value and admin.role != UserRole.super_admin:
        from fastapi import HTTPException as _HTTPException
        raise _HTTPException(403, "Only super admin can create admin users")

    # Only super_admin can set org_id; regular admin always uses their own org
    org_id = admin.org_id
    if data.get("org_id") and admin.role == UserRole.super_admin:
        org_id = data["org_id"]

    new_user = User(
        org_id=org_id,
        email=data["email"],
        hashed_password=hash_password(data["password"]),
        full_name=data["full_name"],
        role=role,
        is_methodist=data.get("is_methodist", False),
    )
    db.add(new_user)
    await db.flush()
    return UserResponse.model_validate(new_user)


# ─── Group Management ─────────────────────────────────────────────────


@router.get("/groups")
async def list_groups_endpoint(
    admin: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
    db: AsyncSession = Depends(get_db),
):
    """List all student groups in the org."""
    from sqlalchemy.orm import selectinload

    from app.admin.models import StudentGroup

    query = select(StudentGroup).options(selectinload(StudentGroup.members))
    if admin.role != UserRole.super_admin:
        query = query.where(StudentGroup.org_id == admin.org_id)
    result = await db.execute(query.order_by(StudentGroup.created_at.desc()))
    groups = result.scalars().unique().all()
    return [
        {
            "id": str(g.id),
            "name": g.name,
            "description": g.description,
            "member_count": len(g.members),
            "created_at": str(g.created_at),
        }
        for g in groups
    ]


@router.post("/groups")
async def create_group_endpoint(
    data: dict,
    admin: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
    db: AsyncSession = Depends(get_db),
):
    """Create a new student group."""
    from app.admin.models import StudentGroup

    group = StudentGroup(
        org_id=admin.org_id,
        name=data["name"],
        description=data.get("description"),
    )
    db.add(group)
    await db.flush()
    return {"id": str(group.id), "name": group.name}


@router.put("/groups/{group_id}")
async def update_group_endpoint(
    group_id: uuid.UUID,
    data: dict,
    admin: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
    db: AsyncSession = Depends(get_db),
):
    """Update group name/description."""
    from app.admin.models import StudentGroup

    query = select(StudentGroup).where(StudentGroup.id == group_id)
    if admin.role != UserRole.super_admin:
        query = query.where(StudentGroup.org_id == admin.org_id)
    result = await db.execute(query)
    group = result.scalar_one_or_none()
    if not group:
        raise NotFoundError("Group not found")

    if "name" in data:
        group.name = data["name"]
    if "description" in data:
        group.description = data["description"]
    await db.flush()
    return {"id": str(group.id), "name": group.name}


@router.delete("/groups/{group_id}")
async def delete_group_endpoint(
    group_id: uuid.UUID,
    admin: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
    db: AsyncSession = Depends(get_db),
):
    """Delete a student group."""
    from app.admin.models import StudentGroup

    query = select(StudentGroup).where(StudentGroup.id == group_id)
    if admin.role != UserRole.super_admin:
        query = query.where(StudentGroup.org_id == admin.org_id)
    result = await db.execute(query)
    group = result.scalar_one_or_none()
    if not group:
        raise NotFoundError("Group not found")
    await db.delete(group)
    return {"ok": True}


@router.get("/groups/{group_id}/members")
async def list_group_members_endpoint(
    group_id: uuid.UUID,
    admin: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
    db: AsyncSession = Depends(get_db),
):
    """List members of a group with user details."""
    from app.admin.models import StudentGroup, StudentGroupMember

    query = select(StudentGroup).where(StudentGroup.id == group_id)
    if admin.role != UserRole.super_admin:
        query = query.where(StudentGroup.org_id == admin.org_id)
    result = await db.execute(query)
    if not result.scalar_one_or_none():
        raise NotFoundError("Group not found")

    result = await db.execute(
        select(User, StudentGroupMember)
        .join(StudentGroupMember, StudentGroupMember.user_id == User.id)
        .where(StudentGroupMember.group_id == group_id)
        .order_by(User.full_name)
    )
    rows = result.all()
    return [
        {
            "id": str(u.id),
            "full_name": u.full_name,
            "email": u.email,
            "role": u.role.value if hasattr(u.role, "value") else u.role,
            "member_id": str(m.id),
        }
        for u, m in rows
    ]


@router.post("/groups/{group_id}/members")
async def add_group_members_endpoint(
    group_id: uuid.UUID,
    data: dict,
    admin: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
    db: AsyncSession = Depends(get_db),
):
    """Add one or more users to a group. data: {user_ids: [uuid, ...]}"""
    from app.admin.models import StudentGroup, StudentGroupMember

    query = select(StudentGroup).where(StudentGroup.id == group_id)
    if admin.role != UserRole.super_admin:
        query = query.where(StudentGroup.org_id == admin.org_id)
    result = await db.execute(query)
    if not result.scalar_one_or_none():
        raise NotFoundError("Group not found")

    user_ids = data.get("user_ids", [])
    added = 0
    for uid in user_ids:
        existing = await db.execute(
            select(StudentGroupMember).where(
                StudentGroupMember.group_id == group_id,
                StudentGroupMember.user_id == uid,
            )
        )
        if existing.scalar_one_or_none():
            continue
        db.add(StudentGroupMember(group_id=group_id, user_id=uid))
        added += 1
    await db.flush()
    return {"added": added}


@router.delete("/groups/{group_id}/members/{user_id}")
async def remove_group_member_endpoint(
    group_id: uuid.UUID,
    user_id: uuid.UUID,
    admin: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
    db: AsyncSession = Depends(get_db),
):
    """Remove a user from a group."""
    from app.admin.models import StudentGroupMember

    result = await db.execute(
        select(StudentGroupMember).where(
            StudentGroupMember.group_id == group_id,
            StudentGroupMember.user_id == user_id,
        )
    )
    member = result.scalar_one_or_none()
    if not member:
        raise NotFoundError("Member not found in group")
    await db.delete(member)
    return {"ok": True}


@router.post("/groups/{group_id}/enroll")
async def enroll_group_endpoint(
    group_id: uuid.UUID,
    data: dict,
    admin: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
    db: AsyncSession = Depends(get_db),
):
    """Enroll all members of a group into a course. data: {course_id: uuid}"""
    from datetime import datetime, timezone

    from app.admin.models import StudentGroup, StudentGroupMember
    from app.courses.models import Course
    from app.progress.models import Enrollment

    query = select(StudentGroup).where(StudentGroup.id == group_id)
    if admin.role != UserRole.super_admin:
        query = query.where(StudentGroup.org_id == admin.org_id)
    result = await db.execute(query)
    if not result.scalar_one_or_none():
        raise NotFoundError("Group not found")

    course_id = data.get("course_id")
    if not course_id:
        from fastapi import HTTPException
        raise HTTPException(400, "course_id required")

    # Cannot enroll into template courses
    course_result = await db.execute(select(Course).where(Course.id == course_id))
    course = course_result.scalar_one_or_none()
    if not course:
        raise NotFoundError("Course not found")
    if getattr(course, "is_template", False):
        from fastapi import HTTPException
        raise HTTPException(400, "Cannot enroll students into a template course. Copy the template first.")

    result = await db.execute(
        select(StudentGroupMember.user_id).where(StudentGroupMember.group_id == group_id)
    )
    member_user_ids = [row[0] for row in result.all()]

    enrolled = 0
    for uid in member_user_ids:
        existing = await db.execute(
            select(Enrollment).where(
                Enrollment.course_id == course_id,
                Enrollment.student_id == uid,
            )
        )
        if existing.scalar_one_or_none():
            continue
        db.add(Enrollment(
            course_id=course_id,
            student_id=uid,
            enrolled_at=datetime.now(timezone.utc),
        ))
        enrolled += 1

    await db.flush()
    return {"enrolled": enrolled, "total_members": len(member_user_ids)}


# ─── Analytics ────────────────────────────────────────────────────────


@router.get("/analytics/export-csv")
async def export_analytics_csv(
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
    db: AsyncSession = Depends(get_db),
):
    """Export enrollment analytics as a CSV file."""
    from app.courses.models import Course
    from app.progress.models import Enrollment

    query = (
        select(User.full_name, User.email, Course.title, Enrollment.progress_percent, Enrollment.enrolled_at, Enrollment.completed_at)
        .join(Enrollment, Enrollment.student_id == User.id)
        .join(Course, Enrollment.course_id == Course.id)
    )
    if user.role != UserRole.super_admin:
        query = query.where(User.org_id == user.org_id)
    result = await db.execute(query.order_by(Enrollment.enrolled_at.desc()))
    rows = result.all()

    def generate_csv():
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["student_name", "student_email", "course_title", "progress_percent", "enrolled_at", "completed_at"])
        for row in rows:
            writer.writerow([
                row.full_name,
                row.email,
                row.title,
                float(row.progress_percent) if row.progress_percent is not None else 0,
                str(row.enrolled_at) if row.enrolled_at else "",
                str(row.completed_at) if row.completed_at else "",
            ])
        output.seek(0)
        yield output.getvalue()

    return StreamingResponse(
        generate_csv(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=analytics_export.csv"},
    )


# ─── Gradebook ─────────────────────────────────────────────────────────


@router.get("/gradebook")
async def gradebook_endpoint(
    course_id: uuid.UUID = Query(...),
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
    db: AsyncSession = Depends(get_db),
):
    """
    Gradebook matrix: rows = enrolled students, columns = assessments.
    Returns { students: [...], columns: [...], rows: { student_id: { col_id: score } } }
    """
    from sqlalchemy import func

    from app.assignments.models import Assignment, AssignmentSubmission
    from app.courses.models import Course, Lesson, Module
    from app.progress.models import Enrollment

    # Verify course access
    course_q = select(Course).where(Course.id == course_id)
    if user.role != UserRole.super_admin:
        course_q = course_q.where(Course.org_id == user.org_id)
    result = await db.execute(course_q)
    course = result.scalar_one_or_none()
    if not course:
        raise NotFoundError("Course not found")

    # Get enrolled students
    result = await db.execute(
        select(User.id, User.full_name, User.email)
        .join(Enrollment, Enrollment.student_id == User.id)
        .where(Enrollment.course_id == course_id)
        .order_by(User.full_name)
    )
    students = [{"id": str(r.id), "full_name": r.full_name, "email": r.email} for r in result.all()]
    student_ids = [s["id"] for s in students]

    if not student_ids:
        return {"students": [], "columns": [], "rows": {}, "averages": {}}

    # Get all lessons in this course (for quizzes, code challenges, interactive)
    result = await db.execute(
        select(Lesson.id)
        .join(Module, Lesson.module_id == Module.id)
        .where(Module.course_id == course_id)
    )
    lesson_ids = [r[0] for r in result.all()]

    columns = []
    rows: dict[str, dict[str, float | None]] = {sid: {} for sid in student_ids}

    # --- Unified Exercises (new system) ---
    if lesson_ids:
        from app.exercises.models import Exercise as Ex
        from app.exercises.models import ExerciseSubmission as ExSub
        result = await db.execute(
            select(Ex).where(Ex.lesson_id.in_(lesson_ids)).order_by(Ex.sort_order)
        )
        exercises = result.scalars().all()
        for ex in exercises:
            col_id = f"exercise_{ex.id}"
            columns.append({
                "id": col_id,
                "type": ex.exercise_type.value if hasattr(ex.exercise_type, 'value') else str(ex.exercise_type),
                "title": ex.title,
                "max_score": 100,
            })
            # Best score per student (score is already 0-100 in exercise_submissions)
            result2 = await db.execute(
                select(ExSub.student_id, func.max(ExSub.score))
                .where(
                    ExSub.exercise_id == ex.id,
                    ExSub.student_id.in_([uuid.UUID(s) for s in student_ids]),
                )
                .group_by(ExSub.student_id)
            )
            for sid, score in result2.all():
                rows[str(sid)][col_id] = round(float(score), 1) if score is not None else None

    # --- Homework Assignments ---
    result = await db.execute(
        select(Assignment).where(Assignment.course_id == course_id)
    )
    assignments = result.scalars().all()
    for a in assignments:
        col_id = f"assignment_{a.id}"
        columns.append({"id": col_id, "type": "assignment", "title": a.title, "max_score": a.max_score})
        result2 = await db.execute(
            select(AssignmentSubmission.student_id, AssignmentSubmission.score)
            .where(
                AssignmentSubmission.assignment_id == a.id,
                AssignmentSubmission.student_id.in_([uuid.UUID(s) for s in student_ids]),
            )
        )
        for sid, score in result2.all():
            if score is not None:
                rows[str(sid)][col_id] = float(score)

    # Compute averages per column
    averages: dict[str, float | None] = {}
    for col in columns:
        cid = col["id"]
        scores = [rows[sid].get(cid) for sid in student_ids if rows[sid].get(cid) is not None]
        averages[cid] = round(sum(scores) / len(scores), 1) if scores else None

    return {"students": students, "columns": columns, "rows": rows, "averages": averages}


@router.get("/gradebook/export")
async def gradebook_export_csv(
    course_id: uuid.UUID = Query(...),
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
    db: AsyncSession = Depends(get_db),
):
    """Export gradebook as CSV."""
    data = await gradebook_endpoint(course_id=course_id, user=user, db=db)

    def generate():
        output = io.StringIO()
        writer = csv.writer(output)
        header = ["Student", "Email"] + [c["title"] for c in data["columns"]] + ["Average"]
        writer.writerow(header)
        for student in data["students"]:
            sid = student["id"]
            scores = []
            for col in data["columns"]:
                val = data["rows"].get(sid, {}).get(col["id"])
                scores.append(val)
            # Compute student average (normalize to 100)
            valid = [s for s in scores if s is not None]
            avg = round(sum(valid) / len(valid), 1) if valid else ""
            writer.writerow(
                [student["full_name"], student["email"]]
                + [s if s is not None else "" for s in scores]
                + [avg]
            )
        # Averages row
        writer.writerow(
            ["AVERAGE", ""]
            + [data["averages"].get(c["id"], "") or "" for c in data["columns"]]
            + [""]
        )
        output.seek(0)
        yield output.getvalue()

    return StreamingResponse(
        generate(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=gradebook_{course_id}.csv"},
    )


@router.get("/gradebook/export-xlsx")
async def gradebook_export_xlsx(
    course_id: uuid.UUID = Query(...),
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
    db: AsyncSession = Depends(get_db),
):
    """Export gradebook as XLSX (Excel) with colour-coded score cells,
    frozen header row, bold averages row, and percent formatting.

    Schools requesting gradebook exports consistently want Excel, not
    CSV — CSV loses formulas, column widths, colours, and everything
    that makes the printed export actually readable. This endpoint
    gives them Excel directly.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    data = await gradebook_endpoint(course_id=course_id, user=user, db=db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Gradebook"

    # Styles
    header_fill = PatternFill("solid", fgColor="4F46E5")  # indigo-600
    header_font = Font(bold=True, color="FFFFFF", size=11)
    avg_row_font = Font(bold=True, italic=True)
    thin = Side(border_style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # Colour thresholds — match the frontend gradebook UI
    green_fill = PatternFill("solid", fgColor="D1FAE5")   # emerald-100
    yellow_fill = PatternFill("solid", fgColor="FEF3C7")  # amber-100
    red_fill = PatternFill("solid", fgColor="FEE2E2")     # rose-100

    def fill_for(score: float | None) -> PatternFill | None:
        if score is None:
            return None
        if score >= 80:
            return green_fill
        if score >= 60:
            return yellow_fill
        return red_fill

    # Header row
    columns = data["columns"]
    header = ["Student", "Email"] + [c["title"] for c in columns] + ["Average"]
    for col_idx, value in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center if col_idx > 2 else left
        cell.border = border

    # Student rows
    row_idx = 2
    for student in data["students"]:
        sid = student["id"]
        ws.cell(row=row_idx, column=1, value=student["full_name"]).border = border
        ws.cell(row=row_idx, column=1).alignment = left
        ws.cell(row=row_idx, column=2, value=student["email"]).border = border
        ws.cell(row=row_idx, column=2).alignment = left

        scores_numeric: list[float] = []
        for i, col in enumerate(columns):
            val = data["rows"].get(sid, {}).get(col["id"])
            cell = ws.cell(row=row_idx, column=3 + i)
            cell.border = border
            cell.alignment = center
            if val is not None:
                cell.value = val
                f = fill_for(val)
                if f is not None:
                    cell.fill = f
                scores_numeric.append(val)
            else:
                cell.value = None

        # Student average
        avg_cell = ws.cell(row=row_idx, column=3 + len(columns))
        avg_cell.border = border
        avg_cell.alignment = center
        if scores_numeric:
            avg = round(sum(scores_numeric) / len(scores_numeric), 1)
            avg_cell.value = avg
            avg_cell.font = Font(bold=True)
            f = fill_for(avg)
            if f is not None:
                avg_cell.fill = f

        row_idx += 1

    # Averages row at the bottom
    avg_row = row_idx
    ws.cell(row=avg_row, column=1, value="Class average").font = avg_row_font
    ws.cell(row=avg_row, column=1).border = border
    ws.cell(row=avg_row, column=1).alignment = left
    ws.cell(row=avg_row, column=2).border = border
    for i, col in enumerate(columns):
        cell = ws.cell(row=avg_row, column=3 + i)
        cell.border = border
        cell.alignment = center
        cell.font = avg_row_font
        v = data["averages"].get(col["id"])
        if v is not None:
            cell.value = v
    ws.cell(row=avg_row, column=3 + len(columns)).border = border

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-size columns (approximate)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 28
    for i in range(len(columns)):
        letter = get_column_letter(3 + i)
        title = columns[i]["title"]
        ws.column_dimensions[letter].width = max(12, min(28, len(title) + 2))
    ws.column_dimensions[get_column_letter(3 + len(columns))].width = 12

    # Stream to memory
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f"attachment; filename=gradebook_{course_id}.xlsx"
        },
    )


# ─── Review Queue ──────────────────────────────────────────────────────


@router.get("/review-queue/count")
async def review_queue_count(
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
    db: AsyncSession = Depends(get_db),
):
    """Count of ungraded assignment submissions for sidebar badge."""
    from sqlalchemy import func

    from app.assignments.models import Assignment, AssignmentStatus, AssignmentSubmission

    query = (
        select(func.count())
        .select_from(AssignmentSubmission)
        .join(Assignment, AssignmentSubmission.assignment_id == Assignment.id)
        .where(
            AssignmentSubmission.status.in_([AssignmentStatus.submitted, AssignmentStatus.late]),
        )
    )
    if user.role != UserRole.super_admin:
        query = query.where(Assignment.org_id == user.org_id)
    if user.role == UserRole.teacher:
        query = query.where(Assignment.created_by == user.id)

    result = await db.execute(query)
    return {"count": result.scalar() or 0}


@router.get("/review-queue")
async def review_queue_list(
    user: User = Depends(require_role(UserRole.admin, UserRole.teacher)),
    db: AsyncSession = Depends(get_db),
):
    """List all ungraded assignment submissions for the teacher/admin."""
    from app.assignments.models import Assignment, AssignmentStatus, AssignmentSubmission

    query = (
        select(AssignmentSubmission, Assignment.title.label("assignment_title"), Assignment.max_score, User.full_name.label("student_name"), User.email.label("student_email"))
        .join(Assignment, AssignmentSubmission.assignment_id == Assignment.id)
        .join(User, AssignmentSubmission.student_id == User.id)
        .where(
            AssignmentSubmission.status.in_([AssignmentStatus.submitted, AssignmentStatus.late]),
        )
    )
    if user.role != UserRole.super_admin:
        query = query.where(Assignment.org_id == user.org_id)
    if user.role == UserRole.teacher:
        query = query.where(Assignment.created_by == user.id)

    result = await db.execute(query.order_by(AssignmentSubmission.submitted_at.asc()))
    rows = result.all()

    return [
        {
            "id": str(sub.id),
            "assignment_id": str(sub.assignment_id),
            "assignment_title": assignment_title,
            "max_score": max_score,
            "student_id": str(sub.student_id),
            "student_name": student_name,
            "student_email": student_email,
            "content": sub.content,
            "file_path": sub.file_path,
            "original_filename": sub.original_filename,
            "submitted_at": str(sub.submitted_at),
            "status": sub.status.value if hasattr(sub.status, "value") else sub.status,
        }
        for sub, assignment_title, max_score, student_name, student_email in rows
    ]
